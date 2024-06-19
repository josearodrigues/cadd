import unicodedata
from openpyxl import load_workbook

def dadosPlanilha():
    wb = load_workbook(filename = 'C:/Users/Eduardo/Dropbox/CEFET/COSAC/ComissaoCADD/11.02.05.99.60-GPROD.xlsx', 
                       use_iterators=True, 
                       read_only=True)
    ws = wb['11.02.05.99.60 -  GPROD']
    
    dados_agregados = {}
    
    print 'Total de linhas = %d' % ws.max_row
    
    qtd_matriculas = 0
    
    for rowNum in range(2, ws.max_row):
        if rowNum % 500 == 0:
            print '# linhas processadas %d' % rowNum
            
        ano = ws.cell(row=rowNum, column=10).value
        if ano in dados_agregados:
            tupla = dados_agregados[ano]
        else:
            tupla = (0, 0, 0)
    
        situacao = ws.cell(row=rowNum, column=12).value
        
        situacao = unicodedata.normalize('NFKD', situacao).encode('ascii','ignore')
        
        if situacao.startswith('Aprovado') or situacao.startswith('Isento') or situacao.startswith('Aproveitamento'):
            aprovacoes = tupla[0] + 1
            reprovacoes = tupla[1]
            trancamentos = tupla[2]
            tupla = (aprovacoes, reprovacoes, trancamentos)
        elif situacao.startswith('Reprovado'):
            aprovacoes = tupla[0]
            reprovacoes = tupla[1] + 1
            trancamentos = tupla[2]
            tupla = (aprovacoes, reprovacoes, trancamentos)
        elif situacao.startswith('Trancamento'):
            aprovacoes = tupla[0]
            reprovacoes = tupla[1]
            trancamentos = tupla[2] + 1
            tupla = (aprovacoes, reprovacoes, trancamentos)
        elif situacao.startswith('Matricula'):
            qtd_matriculas = qtd_matriculas + 1
        else:
            print 'Situacao ignorada: %s' % situacao
            
        dados_agregados[ano] = tupla
    
    print '# matriculas = %d' % qtd_matriculas
            
    for key, value in dados_agregados.iteritems():
        print '(%s, %s)' % key, value
